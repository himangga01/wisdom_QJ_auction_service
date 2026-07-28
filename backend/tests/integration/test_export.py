from io import BytesIO
from datetime import datetime, timezone
from uuid import uuid4

from openpyxl import load_workbook

from app.models import (
    Apartment,
    BrokerArticle,
    BrokerArticleSnapshot,
    ListingGroup,
    TrackedSource,
)
from app.services.export_service import (
    ExportService,
    SHEET_HEADERS,
    create_export_workbook,
    format_korean_won,
    safe_export_filename,
)


def test_write_only_export_has_all_seven_sheets_and_headers() -> None:
    workbook, sheets = create_export_workbook()
    assert workbook.write_only is True
    assert list(sheets) == list(SHEET_HEADERS)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    loaded = load_workbook(output, read_only=True)

    assert loaded.sheetnames == list(SHEET_HEADERS)
    for title, headers in SHEET_HEADERS.items():
        first_row = next(loaded[title].iter_rows(values_only=True))
        assert first_row == tuple(headers)


def test_amount_display_and_filename_are_safe() -> None:
    assert format_korean_won(720_000_000) == "7억 2,000만원"
    assert safe_export_filename("12/34", "20260722-0948") == (
        "naver-land-12-34-20260722-0948.xlsx"
    )


class AsyncRows:
    def __init__(self, rows) -> None:
        self.rows = iter(rows)

    def __aiter__(self):
        return self

    async def __anext__(self):
        try:
            return next(self.rows)
        except StopIteration as error:
            raise StopAsyncIteration from error


class BrokerExportSession:
    def __init__(self, rows) -> None:
        self.rows = rows

    async def stream(self, _statement):
        return AsyncRows(self.rows)


def test_broker_export_marks_disabled_details_and_blanks_detail_json() -> None:
    now = datetime(2026, 7, 28, tzinfo=timezone.utc)
    source = TrackedSource(id=uuid4())
    apartment = Apartment(id=uuid4(), naver_complex_id="12345", name="샘플")
    group = ListingGroup(id=uuid4(), apartment_id=apartment.id)
    article = BrokerArticle(
        id=uuid4(),
        listing_group_id=group.id,
        naver_article_id="1",
        provider="네이버부동산",
        is_npay=False,
        article_url="/articles/1",
        first_seen_at=now,
        last_seen_at=now,
    )
    snapshot = BrokerArticleSnapshot(
        id=uuid4(),
        run_id=uuid4(),
        broker_article_id=article.id,
        details_json={
            "detail_collected": False,
            "market_details": {
                "finance": {"대출": "있음"},
                "transactions": {"거래": "있음"},
                "costs": {"비용": "있음"},
                "maintenance": {"관리": "있음"},
                "complex": {"단지": "있음"},
                "location": {"입지": "있음"},
                "extra_fields": {"기타": "있음"},
            },
        },
        captured_at=now,
    )
    workbook, sheets = create_export_workbook()
    import asyncio

    asyncio.run(
        ExportService(BrokerExportSession([(snapshot, article, group)]))._append_brokers(
            sheets["중개사등록"],
            source,
            apartment,
            None,
            None,
        )
    )
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    row = list(
        load_workbook(output, read_only=True)["중개사등록"].iter_rows(
            values_only=True
        )
    )[1]
    headers = SHEET_HEADERS["중개사등록"]

    assert row[headers.index("추가상세수집여부")] == "N"
    for header in (
        "물건별금융JSON",
        "물건별실거래JSON",
        "물건별비용세금JSON",
        "물건별관리비JSON",
        "물건별단지JSON",
        "물건별입지교통JSON",
        "물건별추가필드JSON",
    ):
        assert row[headers.index(header)] in ("", None)
